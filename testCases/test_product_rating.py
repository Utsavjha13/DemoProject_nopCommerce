from pageObject.login_page import login
from pageObject.fetch_productByRating import find_product_rating
from Utilities.readProperties import ReadConfig
from Utilities.logger import LogGen
import openpyxl


class Test_product_rating:
    App_url = ReadConfig.getAppURL()
    username = ReadConfig.getusername()
    password = ReadConfig.getpassword()
    Create_Logs = LogGen.log_gen()
    file = '/Users/utsav.jha/PycharmProjects/nopCommerce/Download_files/product_ratings.xlsx'
    workbook = openpyxl.load_workbook(file)
    sheet = workbook['Sheet1']
    star = 5

    def test_fetch_product(self, wd):
        self.Create_Logs.info('************* test_fetch_product **********')
        self.driver = wd
        self.driver.maximize_window()
        self.driver.implicitly_wait(10)
        self.driver.get(self.App_url)
        self.lp = login(self.driver)
        self.lp.setUsername(self.username)
        self.lp.setPassword(self.password)
        self.lp.clickLogin()
        self.pr = find_product_rating(self.driver)
        self.pr.product_review_page()
        self.pr.full_page()
        product_list = self.pr.five_star_rating(self.star)
        for i in range(2, len(product_list) + 1):
            self.sheet.cell(i, 1).value = product_list[i - 2]
            self.sheet.cell(i, 2).value = 5
            self.workbook.save(self.file)
        self.Create_Logs.info('************* test_fetch_product>>>PASSED **********')
        assert True

